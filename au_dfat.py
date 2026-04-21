"""Australia DFAT Autonomous Sanctions list adapter.

Source: https://www.dfat.gov.au/sites/default/files/Australian_Sanctions_Consolidated_List.xlsx
Format: XLSX (openpyxl required)
short_code: 'AU_DFAT'

Note: old URL /aut.xlsx is 404. DFAT renamed the file in Nov 2025 to
Australian_Sanctions_Consolidated_List.xlsx. Column structure also updated:
now uses 'Name of Individual or Entity' + 'Name Type' + 'Citizenship'.
"""
from __future__ import annotations

import asyncio
import io
import logging
from typing import Any, Dict, List

from .base import FeedUnavailable, fetch_with_fallback, normalize_entry

logger = logging.getLogger(__name__)

_URLS = [
    # Primary (renamed Nov 2025)
    'https://www.dfat.gov.au/sites/default/files/Australian_Sanctions_Consolidated_List.xlsx',
    # Legacy path (404 since Nov 2025)
    'https://www.dfat.gov.au/sites/default/files/aut.xlsx',
]
_SOURCE_LIST = 'AU_DFAT'
_TIMEOUT = 180.0


class AUDFATAdapter:
    short_code = 'AU_DFAT'

    def __init__(self) -> None:
        self._entries: List[Dict[str, Any]] = []

    async def load(self) -> None:
        raw, used_url = await fetch_with_fallback(
            _URLS, timeout=_TIMEOUT,
            accept='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        # XLSX is a ZIP archive — PK magic validates we got the file, not an HTML error page
        if not raw.startswith(b'PK'):
            raise FeedUnavailable(
                f'AU_DFAT: {used_url} returned non-XLSX bytes (first 4: {raw[:4]!r})'
            )
        self._entries = await asyncio.to_thread(self._parse_xlsx, raw)
        logger.info('AU_DFAT: loaded %d entries from %s', len(self._entries), used_url)

    def get_entries(self) -> List[Dict[str, Any]]:
        return list(self._entries)

    @staticmethod
    def _parse_xlsx(xlsx_bytes: bytes) -> List[Dict[str, Any]]:
        try:
            import openpyxl
        except ImportError:
            logger.error('AU_DFAT: openpyxl is required — pip install openpyxl')
            return []

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active
        out: List[Dict[str, Any]] = []

        # Detect header row — look for the name column substring
        col_idx: Dict[str, int] = {}
        header_row_num = 0
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
            for j, cell in enumerate(row):
                cv = str(cell or '').strip().lower()
                if 'name of individual' in cv or 'name' == cv:
                    # Found header row
                    for k, hdr in enumerate(row):
                        col_idx[str(hdr or '').strip().lower()] = k
                    header_row_num = i + 1
                    break
            if header_row_num:
                break

        def _get(row_vals: tuple, *keys: str) -> str:
            for k in keys:
                for hdr, idx in col_idx.items():
                    if k in hdr and idx < len(row_vals):
                        return str(row_vals[idx] or '').strip()
            return ''

        for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
            if not any(row):
                continue
            row_vals = tuple(row)

            # Skip alias and original-script rows — only emit Primary Name rows
            name_type = _get(row_vals, 'name type')
            if name_type and name_type.lower() not in ('primary name', ''):
                continue

            name = _get(row_vals, 'name of individual', 'name')
            if not name:
                continue

            ref = _get(row_vals, 'reference')
            country = _get(row_vals, 'citizenship', 'nationality', 'country') or None
            if country and len(country) > 50:
                country = None
            program = _get(row_vals, 'committees', 'regime', 'instrument', 'program') or None

            out.append(normalize_entry(
                id=f'AU-{ref or name[:30]}',
                name=name,
                country=country,
                source_list=_SOURCE_LIST,
                aliases=[],
                programs=program or '',
            ))

        return out
